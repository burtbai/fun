# -*- coding: UTF-8 -*-

import json
import os
import pprint
import sys

import openpyxl

in_filename = "cost_in.xlsx"
out_filename = "cost_out.xlsx"


class Cost:
    month = ""
    project_name = ""
    cost_1 = 0
    cost_2 = 0

    def __init__(self, month, project_name, cost_1, cost_2):
        self.month = month
        self.project_name = project_name
        self.cost_1 = cost_1
        self.cost_2 = cost_2

    def __repr__(self):
        return json.dumps(self.__dict__, ensure_ascii=False)

    def update(self, cost_1, cost_2):
        self.cost_1 += cost_1
        self.cost_2 += cost_2


def read_excel(in_filename):
    book = openpyxl.load_workbook(in_filename, data_only=True)

    head_name_month = "月份"
    head_name_project = "项目名称"
    head_name_cost_1 = "含税合计金额"
    head_name_cost_2 = "税额"

    sheet = book.worksheets[0]

    print(in_filename, sheet.title, sheet.dimensions, sheet.max_column, sheet.max_row)

    cost_list = []
    for row in sheet.iter_rows():
        if row[0].row == 1:
            continue

        month = ""
        project_list = []
        cost_1 = ""
        cost_2 = ""
        for cell in row[0: 20]:
            if sheet.cell(row=1, column=cell.column).value == head_name_month:
                month = cell.value
            elif sheet.cell(row=1, column=cell.column).value == head_name_project:
                if cell.value is not None and cell.value != "":
                    project_list = cell.value.split("、")
            elif sheet.cell(row=1, column=cell.column).value == head_name_cost_1:
                cost_1 = cell.value
            elif sheet.cell(row=1, column=cell.column).value == head_name_cost_2:
                cost_2 = cell.value

        for project in project_list:
            exist = False
            for cost in cost_list:
                if month == cost.month and project == cost.project_name:
                    exist = True
                    cost.update(cost_1, cost_2)

            if not exist:
                cost_list.append(Cost(month, project, cost_1, cost_2))

    pprint.pprint(cost_list, width=400)
    return cost_list


def write_excel(cost_list):
    out_wb = openpyxl.Workbook()
    cost_ws_map = {}
    for cost in cost_list:
        if cost.project_name in cost_ws_map:
            cost_ws_map[cost.project_name].append(cost)
        else:
            cost_ws_map[cost.project_name] = [cost]

    head_name = ["费用项目", "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月"]
    for project, costs in cost_ws_map.items():
        out_ws = out_wb.create_sheet(project)
        out_ws.append(head_name)
        out_ws.cell(row=2, column=1).value = "税费-增值税"
        out_ws.cell(row=3, column=1).value = "销售额（含税开票额）"

        for cost in costs:
            for inx, name in enumerate(head_name, start=1):
                if cost.month == name:
                    out_ws.cell(row=2, column=inx).value = cost.cost_2
                    out_ws.cell(row=3, column=inx).value = cost.cost_1

    del out_wb['Sheet']
    out_wb.save(out_filename)


if __name__ == '__main__':
    print("===== 开始 =====")

    path = os.path.dirname(os.path.realpath(sys.executable))
    os.chdir(path)
    print(os.getcwd())

    if not os.path.exists(os.getcwd() + "/" + in_filename):
        print("输入文件名请改成:" + in_filename)
        exit(1)

    data = read_excel(os.getcwd() + "/" + in_filename)
    # data = read_excel(in_filename)
    write_excel(data)

    print("输出文件名:" + out_filename)
    print("===== 完成  =====")
