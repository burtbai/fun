# -*- coding: UTF-8 -*-

import os
import pprint
import sys

import openpyxl

in_path = "/Users/burt/Downloads/test/测试2"
out_filename = "out.xlsx"


def title_match(title_list, key):
    for title in title_list:
        if title is not None and key in str(title):
            return True
    return False


def read_excel(in_filename):
    book = openpyxl.load_workbook(in_filename, data_only=True)

    name_list = ["姓名", "身份证", "应付工资", "失业保险", "养老保险", "医疗保险", "请假工资", "个税", "实发工资",
                 "请假工资", "手机号码"]

    data_map_list = []
    for sheet in book.worksheets:
        print(in_filename, sheet.title, sheet.dimensions, sheet.max_column, sheet.max_row)
        for row in sheet.iter_rows():
            data_map = {}
            for cell in row[0: 40]:
                title_list = list(map(lambda x: sheet.cell(row=x, column=cell.column).value, range(2, 5)))
                if set(title_list) & set(name_list):
                    title = (set(title_list) & set(name_list)).pop()
                    data_map[title] = cell.value
                if title_match(title_list, "公积金"):
                    data_map["公积金"] = cell.value
                if title_match(title_list, "项目"):
                    data_map["项目"] = cell.value
                if title_match(title_list, "身份证"):
                    data_map["身份证"] = cell.value
                if title_match(title_list, "应付工资"):
                    data_map["应付工资"] = cell.value
            if (data_map["姓名"] is not None and 0 < len(data_map["姓名"]) < 5
                    and data_map["姓名"] != "姓名"
                    and data_map["姓名"] != "编制人："):
                if "身份证" not in data_map or data_map["身份证"] is None or len(data_map["身份证"]) < 0:
                    data_map["身份证"] = None
                    print("%s 身份证数据为空" % data_map["姓名"])
                data_map_list.append(data_map)
        print()

    # pprint.pprint(data_map_list, width=400)
    return data_map_list


def save_excel(excel_data_list):
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.worksheets[0]
    out_ws.append(["序号", "姓名", "身份证号码", "手机号码", "应付工资", "子女教育", "继续教育", "房贷利息", "住房租金",
                   "赡养老人", "大病医疗", "养老保险", "医疗保险", "失业保险", "公积金", "个税", "实发工资", "签名",
                   "项目"])
    start = 2
    for excel_data in excel_data_list:
        out_ws.cell(row=start, column=2).value = excel_data["姓名"]
        out_ws.cell(row=start, column=3).value = excel_data["身份证"]
        out_ws.cell(row=start, column=4).value = excel_data["手机号码"]
        out_ws.cell(row=start, column=5).value = excel_data["应付工资"]
        out_ws.cell(row=start, column=12).value = excel_data["养老保险"]
        out_ws.cell(row=start, column=13).value = excel_data["医疗保险"]
        out_ws.cell(row=start, column=14).value = excel_data["失业保险"]
        out_ws.cell(row=start, column=15).value = excel_data["公积金"]
        out_ws.cell(row=start, column=16).value = excel_data["个税"]
        out_ws.cell(row=start, column=17).value = excel_data["实发工资"]
        out_ws.cell(row=start, column=19).value = excel_data["项目"]
        start += 1

    out_wb.save(out_filename)


def handle_data(data_map_list):
    for i in range(len(data_map_list)):
        excel_data = data_map_list[i]
        real = excel_data["应付工资"] if "请假工资" not in excel_data or excel_data["请假工资"] is None \
            else excel_data["应付工资"] - excel_data["请假工资"]
        if "手机号码" not in excel_data:
            excel_data["手机号码"] = None
        excel_data["应付工资"] = real
        if "公积金" not in excel_data:
            excel_data["公积金"] = None
        if "项目" not in excel_data:
            excel_data["项目"] = None
        data_map_list[i] = excel_data

    new_data_map_list = []
    for i in range(len(data_map_list)):
        exist = False
        for j in range(len(new_data_map_list)):
            if (data_map_list[i]["姓名"] == new_data_map_list[j]["姓名"]
                    and data_map_list[i]["身份证"] is not None
                    and data_map_list[i]["身份证"] == new_data_map_list[j]["身份证"]):
                exist = True
                new_data_map_list[j]["应付工资"] += data_map_list[i]["应付工资"]
                if new_data_map_list[j]["养老保险"] is not None:
                    new_data_map_list[j]["养老保险"] += data_map_list[i]["养老保险"]
                if new_data_map_list[j]["医疗保险"] is not None:
                    new_data_map_list[j]["医疗保险"] += data_map_list[i]["医疗保险"]
                if new_data_map_list[j]["失业保险"] is not None:
                    new_data_map_list[j]["失业保险"] += data_map_list[i]["失业保险"]
                if new_data_map_list[j]["公积金"] is not None:
                    new_data_map_list[j]["公积金"] += data_map_list[i]["公积金"]
                if new_data_map_list[j]["个税"] is not None:
                    new_data_map_list[j]["个税"] += data_map_list[i]["个税"]
                new_data_map_list[j]["实发工资"] += data_map_list[i]["实发工资"]
                print("%s + %s 相同, 数据合并" % (data_map_list[i]["姓名"], data_map_list[i]["身份证"]))
        if not exist:
            new_data_map_list.append(data_map_list[i])

    return new_data_map_list


if __name__ == '__main__':
    path = os.path.dirname(os.path.realpath(sys.executable))
    # path = in_path  # 测试
    os.chdir(path)
    print(os.getcwd())

    files = list()
    for item in os.scandir():
        if (item.is_file() and item.name.endswith(".xlsx") and "~$" not in item.name
                and item.name != out_filename and ".~" not in item.name):
            files.append(item.path)
    files.sort()
    print(files)

    data_list = list()
    for file in files:
        data = read_excel(file)
        data_list += data

    data_list = handle_data(data_list)
    save_excel(data_list)

    print("===== 完成 =====")
